# -*- coding: utf-8 -*-
"""
DRY-RUN Fase A (READ-ONLY, no toca nada): clientes FACTURADOS jul-2022->hoy en Tactica
vs CRM Neon, separando transportes. Reusa el modelo de identidad de _tactica_export_compradores.py.
"""
import io, os, re, json, sys
sys.stdout.reconfigure(encoding='utf-8')

SQL = r"D:\Dropbox\FEBECOS - ADMINISTRACION\backup-tactica\tactica.sql"
XLSX = r"D:\Dropbox\FEBECOS - FULL CLAUDE\FEBO Administración\exports\tactica_contactos_y_transportes.xlsx"
SNAP = r"C:\Users\Guille\AppData\Local\Temp\claude\D--Dropbox-FEBECOS---FULL-CLAUDE\e0607c3a-1d61-4664-aec4-3ddff761a026\scratchpad\crm-snapshot.json"
OUT = r"C:\Users\Guille\AppData\Local\Temp\claude\D--Dropbox-FEBECOS---FULL-CLAUDE\e0607c3a-1d61-4664-aec4-3ddff761a026\scratchpad\dryrun_fase_a.json"
CUTOFF = '2022-07-01'

# ---- parser del dump (copiado de _tactica_export_compradores.py) ----
def parse_row(s):
    vals=[];i=0;n=len(s)
    while i<n:
        while i<n and s[i] in ' \t': i+=1
        if i>=n: break
        if s[i]=="'":
            i+=1;buf=[]
            while i<n:
                c=s[i]
                if c=='\\':
                    if i+1<n: buf.append({'n':'\n','t':'\t','r':'\r','0':'\0'}.get(s[i+1],s[i+1]));i+=2;continue
                    else: i+=1;continue
                if c=="'":
                    if i+1<n and s[i+1]=="'": buf.append("'");i+=2;continue
                    i+=1;break
                buf.append(c);i+=1
            vals.append(''.join(buf))
        else:
            j=i
            while j<n and s[j]!=',': j+=1
            tok=s[i:j].strip();vals.append(None if tok.upper()=='NULL' else tok);i=j
        while i<n and s[i] in ' \t': i+=1
        if i<n and s[i]==',': i+=1
    return vals

def iter_rows(table):
    prefix="INSERT INTO `%s` VALUES "%table
    with io.open(SQL,'r',encoding='utf-8',errors='replace') as f:
        for line in f:
            if not line.startswith(prefix): continue
            body=line[len(prefix):].rstrip()
            if body.endswith(';'): body=body[:-1]
            i=0;n=len(body)
            while i<n:
                if body[i]!='(': i+=1;continue
                i+=1;start=i;in_str=False
                while i<n:
                    c=body[i]
                    if in_str:
                        if c=='\\': i+=2;continue
                        if c=="'":
                            if i+1<n and body[i+1]=="'": i+=2;continue
                            in_str=False;i+=1;continue
                        i+=1
                    else:
                        if c=="'": in_str=True;i+=1
                        elif c==')':
                            yield parse_row(body[start:i]);i+=1;break
                        else: i+=1

def g(r, idx): return r[idx] if idx < len(r) else None
def clean(v): return (v or '').strip()
def norm_cuit(v):
    d = re.sub(r'\D', '', v or '')
    return d if len(d) == 11 else None
def norm_tel(v):
    d = re.sub(r'\D', '', v or '')
    return d[-10:] if len(d) >= 8 else None
def norm_email(v):
    v = (v or '').strip().lower()
    return v if re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', v) else None
def norm_name(v):
    return re.sub(r'[^a-z0-9]', '', (v or '').lower())

# ---- cargar dump ----
print("Cargando empresas/contactos/fiscal/telefonos/emails del dump...")
EMP = dict(RecID=0, IDEmpresa=1, Empresa=2, Bloqueado=7, Eliminado=8)
empresas = {}
for r in iter_rows('empresas'):
    if g(r, EMP['Eliminado']) == '1' or g(r, EMP['Bloqueado']) == '1': continue
    idemp = g(r, EMP['IDEmpresa'])
    if idemp: empresas[idemp] = dict(Empresa=clean(g(r, EMP['Empresa'])))

CON = dict(RecID=0, IDContacto=1, IDEmpresa=2, Nombre=5, Apellido=12, Bloqueado=11, Eliminado=16, Correo=29, Documento=30)
contactos = {}
for r in iter_rows('contactos'):
    if g(r, CON['Eliminado']) == '1' or g(r, CON['Bloqueado']) == '1': continue
    idc = g(r, CON['IDContacto'])
    if idc: contactos[idc] = dict(RecID=g(r, CON['RecID']), IDEmpresa=g(r, CON['IDEmpresa']),
        Nombre=clean(g(r, CON['Nombre'])), Apellido=clean(g(r, CON['Apellido'])), Correo=clean(g(r, CON['Correo'])))

FIS = dict(RecID=0, RazonSocial=1, IDRef=2, NroImp1=9, Bloqueado=30)
fiscal_by_recid = {}
for r in iter_rows('fiscal'):
    if g(r, FIS['Bloqueado']) == '1': continue
    fiscal_by_recid[g(r, FIS['RecID'])] = dict(RazonSocial=clean(g(r, FIS['RazonSocial'])),
        IDRef=g(r, FIS['IDRef']), CUIT=norm_cuit(g(r, FIS['NroImp1'])))

DIR = dict(RecID=0, IDRef=1)
dir_recid_to_ref = {g(r, 0): g(r, 1) for r in iter_rows('direcciones')}

tel_by_emp, tel_by_con = {}, {}
def add_tel(d, key, num):
    if not num: return
    lst = d.setdefault(key, [])
    if num not in lst: lst.append(num)
for r in iter_rows('telefonos'):
    num = clean(g(r, 2))
    if not num: continue
    ref = dir_recid_to_ref.get(g(r, 1)); ref2 = g(r, 5)
    if ref in empresas: add_tel(tel_by_emp, ref, num)
    elif ref in contactos: add_tel(tel_by_con, ref, num)
    if ref2 and ref2 in contactos: add_tel(tel_by_con, ref2, num)

mail_by_con = {}
for r in iter_rows('direccionescorreo'):
    em = clean(g(r, 2)); ref = g(r, 1)
    if em and ref in contactos:
        lst = mail_by_con.setdefault(ref, [])
        if em not in lst: lst.append(em)
for idc, c in contactos.items():
    if c['Correo']:
        lst = mail_by_con.setdefault(idc, [])
        if c['Correo'] not in lst: lst.append(c['Correo'])

con_by_emp = {}
for idc, c in contactos.items():
    if c['IDEmpresa']: con_by_emp.setdefault(c['IDEmpresa'], []).append(idc)

def emp_tels(idemp):
    nums = list(tel_by_emp.get(idemp, []))
    for idc in con_by_emp.get(idemp, []):
        for n in tel_by_con.get(idc, []):
            if n not in nums: nums.append(n)
    return nums
def emp_mails(idemp):
    mails = []
    for idc in con_by_emp.get(idemp, []):
        for m in mail_by_con.get(idc, []):
            if m not in mails: mails.append(m)
    return mails

# ---- facturados jul-2022 -> hoy ----
print("Cargando facturas...")
FE = dict(RecID=0, Estado=5, FechaEmision=11, IDFiscal=13, Tipo=32)
VAL = {'0', '1', '5'}
last, first, cant = {}, {}, {}
for r in iter_rows('facturas'):
    if len(r) <= FE['Tipo']: continue
    if g(r, FE['Tipo']) != '0' or g(r, FE['Estado']) not in VAL: continue
    fe = g(r, FE['FechaEmision'])
    if not fe or len(fe) < 10 or fe[:10] < CUTOFF: continue
    idfis = g(r, FE['IDFiscal'])
    if not idfis: continue
    cant[idfis] = cant.get(idfis, 0) + 1
    if idfis not in last or fe > last[idfis]: last[idfis] = fe
print(f"fiscal RecIDs facturados >= {CUTOFF}: {len(last)}")

# ---- transportes del xlsx (para excluir del CRM) ----
import openpyxl
wb = openpyxl.load_workbook(XLSX, read_only=True)
transp_rows = []
ws = wb['Transportes']
headers = [c.value for c in next(ws.iter_rows(max_row=1))]
for row in ws.iter_rows(min_row=2):
    d = dict(zip(headers, [c.value for c in row]))
    if d.get('Nombre'): transp_rows.append(d)
transp_names = {norm_name(t['Nombre']) for t in transp_rows}
transp_tels = {norm_tel(str(t.get('Teléfono principal') or '')) for t in transp_rows} - {None}
print(f"transportes en xlsx: {len(transp_rows)}")

# ---- CRM snapshot ----
snap = json.load(io.open(SNAP, encoding='utf-8'))
crm = snap['clientes']; carriers = snap['carriers']
crm_by_cuit, crm_by_mail, crm_by_tel = {}, {}, {}
for c in crm:
    cu = norm_cuit(c.get('cuit'));  em = norm_email(c.get('email')); te = norm_tel(c.get('whatsapp'))
    if cu: crm_by_cuit.setdefault(cu, c)
    if em: crm_by_mail.setdefault(em, c)
    if te: crm_by_tel.setdefault(te, c)
carrier_names = {norm_name(x['name']) for x in carriers}

# ---- clasificar facturados ----
res = dict(total=0, transportes_excluidos=0, con_email=0, sin_email=0,
           ya_en_crm=0, nuevos=0, ya_cliente_final=0, ya_revendedor=0,
           a_marcar_cliente_final=0, a_tag_compro_rev=0, nuevos_con_email=0, nuevos_sin_email=0)
detalle_nuevos, detalle_upgrade, detalle_rev = [], [], []
for idfis, fecha in last.items():
    fis = fiscal_by_recid.get(idfis)
    if not fis: continue
    ref = fis['IDRef']; nombre = fis['RazonSocial']; cuitn = fis['CUIT']
    tels, mails = [], []
    if ref in empresas:
        nombre = empresas[ref]['Empresa'] or nombre
        tels, mails = emp_tels(ref), emp_mails(ref)
    elif ref in contactos:
        c = contactos[ref]
        nombre = (c['Nombre'] + ' ' + c['Apellido']).strip() or nombre
        tels = list(tel_by_con.get(c['RecID'], [])); mails = list(mail_by_con.get(ref, []))
    # transporte?
    if norm_name(nombre) in transp_names or norm_name(nombre) in carrier_names or any(norm_tel(t) in transp_tels for t in tels):
        res['transportes_excluidos'] += 1
        continue
    res['total'] += 1
    emails_ok = [m for m in (norm_email(m) for m in mails) if m]
    res['con_email' if emails_ok else 'sin_email'] += 1
    # dedup CRM: cuit -> email -> tel
    match = (cuitn and crm_by_cuit.get(cuitn)) or next((crm_by_mail[m] for m in emails_ok if m in crm_by_mail), None) \
        or next((crm_by_tel[t] for t in (norm_tel(t) for t in tels) if t and t in crm_by_tel), None)
    if match:
        res['ya_en_crm'] += 1
        t = match.get('tipo')
        if t == 'cliente_final': res['ya_cliente_final'] += 1
        elif t == 'revendedor':
            res['ya_revendedor'] += 1; res['a_tag_compro_rev'] += 1
            detalle_rev.append(dict(crm_id=match['id'], nombre=nombre))
        else:
            res['a_marcar_cliente_final'] += 1
            detalle_upgrade.append(dict(crm_id=match['id'], nombre=nombre, tipo_actual=t, cuit=cuitn, emails=emails_ok, ultima=fecha[:10], facturas=cant.get(idfis)))
    else:
        res['nuevos'] += 1
        res['nuevos_con_email' if emails_ok else 'nuevos_sin_email'] += 1
        detalle_nuevos.append(dict(nombre=nombre, cuit=cuitn, emails=emails_ok, tels=tels[:2], ultima=fecha[:10], facturas=cant.get(idfis)))

# ---- transportes xlsx vs carriers ----
tr = dict(total=len(transp_rows), con_email=0, ya_en_carriers=0, nuevos=0)
for t in transp_rows:
    if norm_email(str(t.get('Email') or '')): tr['con_email'] += 1
    if norm_name(t['Nombre']) in carrier_names: tr['ya_en_carriers'] += 1
    else: tr['nuevos'] += 1

print('\n=== RESULTADO CLIENTES FACTURADOS jul-2022→hoy ===')
for k, v in res.items(): print(f'  {k}: {v}')
print('\n=== TRANSPORTES (xlsx vs logistics.carriers, ya hay', len(carriers), ') ===')
for k, v in tr.items(): print(f'  {k}: {v}')
json.dump(dict(resumen=res, transportes=tr, nuevos=detalle_nuevos, upgrades=detalle_upgrade, revendedores=detalle_rev),
          io.open(OUT, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
print('\nDetalle nominal →', OUT)
